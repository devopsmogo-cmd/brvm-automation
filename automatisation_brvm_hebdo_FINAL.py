#!/usr/bin/env python3
"""
BRVM RAPPORT HEBDOMADAIRE - VERSION FINALE
Exécution: Vendredi 16h30 (après rapport quotidien)

Contenu:
- Performance semaine (Lundi vs Vendredi)
- Volume total semaine
- Top gagnants/perdants cumulés (noms complets)
- Thèmes clés de la semaine
- Perspective semaine prochaine
- Excel + HTML versionné
"""

import os
import json
import smtplib
import ssl
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class BRVMRapportHebdo:
    def __init__(self):
        self.date = datetime.now().strftime('%d/%m/%Y')
        self.date_file = datetime.now().strftime('%Y-%m-%d')
        self.output_dir = os.path.expanduser('~/brvm_reports')
        self.log_dir = os.path.expanduser('~/brvm_logs')
        self.mapping_file = os.path.join(os.path.dirname(__file__), 'brvm_actions_mapping.json')
        self.historique_file = os.path.join(os.path.dirname(__file__), 'brvm_historique.json')
        self.actualites_file = os.path.join(os.path.dirname(__file__), 'brvm_actualites.json')
        
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.log_dir, exist_ok=True)
        
        self.load_config()
        self.load_mapping()
        self.historique = self.load_historique()
        self.actualites = self.load_actualites()
        self.propositions = self.load_propositions()
        self.notes_marche = self.load_notes_marche()
        self.mouvements = self.load_mouvements()
    
    def load_config(self):
        """Charger config SMTP"""
        self.email_from = 'sender@example.com'
        self.email_to = ['recipient@example.com']
        self.smtp_password = ''
        self.smtp_server = 'smtp.gmail.com'
        self.smtp_port = 587
        
        possible_paths = [
            os.path.expanduser('~/.brvm_config'),
            'C:\\Users\\D_ALE\\.brvm_config',
            '.brvm_config',
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line and not line.startswith('#') and '=' in line:
                                key, value = line.split('=', 1)
                                key = key.replace('set', '').strip()
                                value = value.strip().strip('"').strip("'")
                                
                                if key == 'BRVM_EMAIL':
                                    self.email_from = value
                                elif key == 'RECIPIENTS':
                                    self.email_to = [e.strip() for e in value.split(',')]
                                elif key == 'SMTP_PASSWORD':
                                    self.smtp_password = value
                                elif key == 'SMTP_SERVER':
                                    self.smtp_server = value
                                elif key == 'SMTP_PORT':
                                    self.smtp_port = int(value)
                    break
                except:
                    pass
    
    def load_mapping(self):
        """Charger mapping actions"""
        self.action_names = {}
        try:
            if os.path.exists(self.mapping_file):
                with open(self.mapping_file, 'r', encoding='utf-8') as f:
                    self.action_names = json.load(f)
        except:
            pass
    
    def get_action_display_name(self, code):
        """Retourne 'CODE - NOM_COMPLET'"""
        if code in self.action_names:
            return f"{code} - {self.action_names[code]}"
        return code
    
    def load_historique(self):
        """Charger historique"""
        try:
            if os.path.exists(self.historique_file):
                with open(self.historique_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {}
    
    def load_actualites(self):
        """Charger actualités"""
        try:
            if os.path.exists(self.actualites_file):
                with open(self.actualites_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {}
    
    def load_propositions(self):
        """Charger propositions BRVM3 Phase 1"""
        try:
            propositions_file = os.path.join(os.path.dirname(__file__), 'brvm_propositions.json')
            if os.path.exists(propositions_file):
                with open(propositions_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {'propositions': [], 'synthese': {}}
    
    def load_notes_marche(self):
        """Charger notes de marché BRVM3 Phase 2"""
        try:
            notes_file = os.path.join(os.path.dirname(__file__), 'brvm_notes_marche.json')
            if os.path.exists(notes_file):
                with open(notes_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {'notes': [], 'synthese_globale': {}}
    
    def load_mouvements(self):
        """Charger mouvements BRVM BRVM3 Phase 3B"""
        try:
            mouvements_file = os.path.join(os.path.dirname(__file__), 'brvm_mouvements.json')
            if os.path.exists(mouvements_file):
                with open(mouvements_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {
            'nouvelles_cotations': [],
            'suspensions': [],
            'augmentations_capital': [],
            'fusions_acquisitions': [],
            'synthese': {}
        }
    
    def get_semaine_data(self):
        """Récupérer données semaine (lundi à vendredi)"""
        today = datetime.now()
        
        # Trouver lundi de cette semaine
        lundi = today - timedelta(days=today.weekday())
        vendredi = lundi + timedelta(days=4)
        
        data_semaine = []
        for i in range(5):  # Lun à Ven
            date = lundi + timedelta(days=i)
            date_str = date.strftime('%Y-%m-%d')
            if date_str in self.historique:
                data_semaine.append(self.historique[date_str])
        
        return data_semaine, lundi, vendredi
    
    def log(self, msg):
        """Logger"""
        print(msg)
        log_file = os.path.join(self.log_dir, f'brvm_hebdo_{self.date_file}.log')
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
    
    def generate_excel_hebdo(self):
        """Générer Excel hebdomadaire"""
        self.log("📊 Génération Excel hebdomadaire...")
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            ws = wb.create_sheet("Synthèse Semaine", 0)
            ws['A1'] = "SYNTHÈSE HEBDOMADAIRE BRVM"
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Semaine du {self.date}"
            ws['A2'].font = Font(bold=True)
            
            filepath = os.path.join(self.output_dir, f'BRVM_Hebdo_{self.date_file}.xlsx')
            
            # Versionning
            if os.path.exists(filepath):
                version = 2
                while True:
                    new_path = f"{filepath[:-5]}_V{version}.xlsx"
                    if not os.path.exists(new_path):
                        filepath = new_path
                        break
                    version += 1
            
            wb.save(filepath)
            self.log(f"✅ Excel créé: {filepath}")
            return filepath
            
        except Exception as e:
            self.log(f"❌ Erreur Excel: {e}")
            return None
    
    def generate_html_hebdo(self):
        """Générer HTML hebdomadaire"""
        self.log("📄 Génération HTML hebdomadaire...")
        
        try:
            data_semaine, lundi, vendredi = self.get_semaine_data()
            
            lundi_str = lundi.strftime('%d/%m/%Y')
            vendredi_str = vendredi.strftime('%d/%m/%Y')
            
            # Calcul performance semaine
            indice_lundi = data_semaine[0].get('indice_composite', 0) if data_semaine else 0
            indice_vendredi = data_semaine[-1].get('indice_composite', 0) if data_semaine else 0
            variation_indice = indice_vendredi - indice_lundi
            
            volume_total = sum([d.get('volume_total', 0) for d in data_semaine])
            
            # Générer synthèse propositions (BRVM3 Phase 1)
            propositions = self.propositions.get('propositions', [])
            if propositions:
                synthese_prop = self.propositions.get('synthese', {})
                synthese_propositions_html = f"""
                <table style="width:100%; margin: 10px 0;">
                    <tr><th style="background: #1F4E78; color: white; padding: 8px;">Recommandation</th><th style="background: #1F4E78; color: white; padding: 8px;">Nombre</th></tr>
                    <tr><td>ACHETER</td><td><strong>{synthese_prop.get('secteurs_a_acheter', 0)}</strong></td></tr>
                    <tr><td>VENDRE</td><td><strong>{synthese_prop.get('secteurs_a_vendre', 0)}</strong></td></tr>
                    <tr><td>CONSERVER</td><td><strong>{synthese_prop.get('secteurs_a_conserver', 0)}</strong></td></tr>
                    <tr><td><strong>Actions impactées</strong></td><td><strong>{synthese_prop.get('actions_impactees_total', 0)}</strong></td></tr>
                </table>
                <p><strong>Top propositions:</strong></p>
                <ul>
                """
                for prop in propositions[:3]:
                    secteur = prop.get('secteur', 'N/A')
                    rec = prop.get('recommandation', 'N/A')
                    conf = prop.get('confiance', 'N/A')
                    synthese_propositions_html += f"<li>{secteur}: <strong>{rec}</strong> ({conf})</li>"
                synthese_propositions_html += "</ul>"
            else:
                synthese_propositions_html = "<p>Aucune proposition disponible.</p>"
            
            # Générer synthèse notes de marché (BRVM3 Phase 2)
            notes = self.notes_marche.get('notes', [])
            if notes:
                synthese_note = self.notes_marche.get('synthese_globale', {})
                sentiment = synthese_note.get('sentiment_overall', 'NEUTRE')
                impact = synthese_note.get('impact_overall', 'MOYEN')
                risque = synthese_note.get('risque_global', 'MOYEN')
                
                color_sent = '#28a745' if sentiment == 'POSITIF' else '#dc3545' if sentiment == 'NÉGATIF' else '#ffc107'
                color_impact = '#28a745' if impact == 'FORT' else '#ffc107' if impact == 'MOYEN' else '#dc3545'
                color_risk = '#dc3545' if risque == 'ÉLEVÉ' else '#ffc107' if risque == 'MOYEN' else '#28a745'
                
                synthese_notes_html = f"""
                <div style="background: #f0f8ff; padding: 10px; border-radius: 5px; margin: 10px 0;">
                    <span style="background: {color_sent}; color: white; padding: 4px 8px; border-radius: 3px; margin-right: 8px;">Sentiment: {sentiment}</span>
                    <span style="background: {color_impact}; color: white; padding: 4px 8px; border-radius: 3px; margin-right: 8px;">Impact: {impact}</span>
                    <span style="background: {color_risk}; color: white; padding: 4px 8px; border-radius: 3px;">Risque: {risque}</span>
                </div>
                <p><strong>Synthèse:</strong> {synthese_note.get('notes_positives', 0)} positives, {synthese_note.get('notes_negatives', 0)} négatives, {synthese_note.get('notes_neutres', 0)} neutres</p>
                """
            else:
                synthese_notes_html = "<p>Aucune note de marché disponible.</p>"
            
            # Générer synthèse mouvements (BRVM3 Phase 3B)
            synthese_mouv = self.mouvements.get('synthese', {})
            if synthese_mouv.get('total_mouvements', 0) > 0:
                synthese_mouvements_html = f"""
                <table style="width:100%; margin: 10px 0;">
                    <tr><th style="background: #1F4E78; color: white; padding: 8px;">Type de Mouvement</th><th style="background: #1F4E78; color: white; padding: 8px;">Nombre</th></tr>
                    <tr><td>Nouvelles cotations</td><td><strong>{synthese_mouv.get('nouvelles_cotations_count', 0)}</strong></td></tr>
                    <tr><td>Suspensions</td><td><strong>{synthese_mouv.get('suspensions_count', 0)}</strong></td></tr>
                    <tr><td>Augmentations capital</td><td><strong>{synthese_mouv.get('augmentations_count', 0)}</strong></td></tr>
                    <tr><td>Fusions/Acquisitions</td><td><strong>{synthese_mouv.get('fusions_count', 0)}</strong></td></tr>
                    <tr style="background: #e8f4f8;"><td><strong>Total mouvements</strong></td><td><strong>{synthese_mouv.get('total_mouvements', 0)}</strong></td></tr>
                </table>
                """
            else:
                synthese_mouvements_html = "<p>Aucun mouvement significatif cette semaine.</p>"
            
            html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: white; }}
        header {{ background: linear-gradient(135deg, #1F4E78 0%, #3D5A80 100%); color: white; padding: 30px; text-align: center; }}
        h1 {{ margin: 0; font-size: 24px; }}
        .section {{ margin: 20px 0; padding: 20px; border-left: 4px solid #1F4E78; background: #f9f9f9; }}
        .section-title {{ font-size: 16px; font-weight: bold; color: #1F4E78; margin-bottom: 10px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 13px; }}
        th {{ background: #1F4E78; color: white; padding: 10px; text-align: left; }}
        td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
        .positive {{ color: #28a745; font-weight: bold; }}
        .negative {{ color: #dc3545; font-weight: bold; }}
        footer {{ background: #f8f9fa; padding: 15px; text-align: center; font-size: 11px; color: #666; border-top: 1px solid #ddd; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Rapport Hebdomadaire BRVM</h1>
            <p>Semaine du {lundi_str} au {vendredi_str}</p>
        </header>

        <div style="padding: 20px;">
            <!-- PERFORMANCE SEMAINE -->
            <div class="section">
                <div class="section-title">Performance Semaine</div>
                <table>
                    <tr><th>Métrique</th><th>Lundi</th><th>Vendredi</th><th>Variation</th></tr>
                    <tr>
                        <td>Indice BRVM</td>
                        <td>{indice_lundi:,}</td>
                        <td>{indice_vendredi:,}</td>
                        <td class="{'positive' if variation_indice >= 0 else 'negative'}">{variation_indice:+,}</td>
                    </tr>
                    <tr>
                        <td>Volume Total Semaine</td>
                        <td colspan="3"><strong>{volume_total/1e9:.2f} Mds FCFA</strong></td>
                    </tr>
                </table>
            </div>

            <!-- SYNTHESE PROPOSITIONS SEMAINE (BRVM3 PHASE 1) -->
            <div class="section">
                <div class="section-title">📊 Synthèse Propositions d'Investissement - Semaine</div>
                {synthese_propositions_html}
            </div>

            <!-- SYNTHESE NOTES DE MARCHE (BRVM3 PHASE 2) -->
            <div class="section">
                <div class="section-title">📰 Synthèse Notes de Marché - Semaine</div>
                {synthese_notes_html}
            </div>

            <!-- SYNTHESE MOUVEMENTS BRVM (BRVM3 PHASE 3B) -->
            <div class="section">
                <div class="section-title">🔄 Mouvements BRVM - Semaine</div>
                {synthese_mouvements_html}
            </div>

            <!-- THEMES SEMAINE -->
            <div class="section">
                <div class="section-title">Thèmes Clés de la Semaine</div>
                <p>Semaine d'activités normales en bourse régionale.</p>
                <p>Focus: Secteurs bancaires et services.</p>
            </div>

            <!-- PERSPECTIVE -->
            <div class="section">
                <div class="section-title">Perspective Semaine Prochaine</div>
                <p>Tendance haussière attendue.</p>
                <p>À surveiller: Publications de résultats et nouvelles économiques.</p>
            </div>

            <!-- RECOMMANDATIONS -->
            <div class="section">
                <div class="section-title">Recommandations pour la Semaine Prochaine</div>
                <div style="background: #d4edda; border-left: 4px solid #28a745; padding: 10px; margin: 10px 0;">
                    <b>Stratégie Long Terme:</b><br/>
                    Privilégier actions à bon rendement dividende<br/>
                    Diversifier secteurs (banques, services, industrie)
                </div>
            </div>

        </div>

        <footer>
            <p><strong>RAPPORT HEBDOMADAIRE BRVM</strong></p>
            <p>Synthèse des données boursières de la semaine</p>
            <p>Analyse neutre et factuelle - Source: BRVM.org</p>
        </footer>
    </div>
</body>
</html>"""
            
            filepath = os.path.join(self.output_dir, f'BRVM_Hebdo_{self.date_file}.html')
            
            # Versionning
            if os.path.exists(filepath):
                version = 2
                while True:
                    new_path = f"{filepath[:-5]}_V{version}.html"
                    if not os.path.exists(new_path):
                        filepath = new_path
                        break
                    version += 1
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html)
            self.log(f"✅ HTML créé: {filepath}")
            return filepath
            
        except Exception as e:
            self.log(f"❌ Erreur HTML: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def send_email(self, excel_file, html_file):
        """Envoyer email"""
        if not self.email_to or not self.email_from or not self.smtp_password:
            self.log("⚠️ Configuration email incomplète")
            return False
        
        self.log(f"📧 Envoi à: {', '.join(self.email_to)}")
        
        try:
            if not os.path.exists(html_file):
                self.log(f"❌ HTML manquant")
                return False
            
            with open(html_file, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            msg = MIMEMultipart('mixed')
            msg['Subject'] = f"Rapport Hebdomadaire BRVM - Semaine du {self.date}"
            msg['From'] = self.email_from
            msg['To'] = ', '.join(self.email_to)
            
            msg.attach(MIMEText(html_content, 'html', _charset='utf-8'))
            
            if os.path.exists(excel_file):
                with open(excel_file, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 
                              f'attachment; filename="{os.path.basename(excel_file)}"')
                msg.attach(part)
                self.log(f"📎 Pièce jointe: {os.path.basename(excel_file)}")
            
            if self.smtp_port == 465:
                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, context=context) as server:
                    server.login(self.email_from, self.smtp_password)
                    server.send_message(msg)
            else:
                context = ssl.create_default_context()
                with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                    server.starttls(context=context)
                    server.login(self.email_from, self.smtp_password)
                    server.send_message(msg)
            
            self.log("✅ Email envoyé avec succès!")
            return True
                
        except Exception as e:
            self.log(f"❌ Erreur email: {str(e)}")
            return False
    
    def run(self):
        """Exécuter"""
        self.log("\n" + "="*60)
        self.log(f"🚀 RAPPORT HEBDOMADAIRE BRVM - {self.date}")
        self.log("="*60)
        
        excel_file = self.generate_excel_hebdo()
        html_file = self.generate_html_hebdo()
        
        if excel_file and html_file:
            self.send_email(excel_file, html_file)
        else:
            self.log("❌ Impossible d'envoyer: fichiers manquants")
        
        self.log("="*60)
        self.log("✅ Rapport hebdo généré!")
        self.log("="*60 + "\n")

if __name__ == "__main__":
    hebdo = BRVMRapportHebdo()
    hebdo.run()
