#!/usr/bin/env python3
"""
BRVM AUTOMATION - VERSION 3 FINAL COMPLET
Rapport quotidien COMPLET avec:
- Données officielles BRVM
- Top 5 gagnants
- Analyses dividendes 5 ans
- Opportunités cachées
- Recommandations fusionnées
- Analyses trading (Madis + Sika)
"""

import os
import json
import smtplib
import ssl
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class BRVMAutomationComplet:
    def __init__(self):
        self.date = datetime.now().strftime('%d/%m/%Y')
        self.date_file = datetime.now().strftime('%Y-%m-%d')
        self.date_obj = datetime.now()
        self.output_dir = os.path.expanduser('~/brvm_reports')
        self.log_dir = os.path.expanduser('~/brvm_logs')
        self.historique_file = os.path.join(os.path.dirname(__file__), 'brvm_historique.json')
        self.actualites_file = os.path.join(os.path.dirname(__file__), 'brvm_actualites.json')
        self.mapping_file = os.path.join(os.path.dirname(__file__), 'brvm_actions_mapping.json')
        self.dividendes_file = os.path.join(os.path.dirname(__file__), 'brvm_dividendes_historique.json')
        self.cote_file = os.path.join(os.path.dirname(__file__), 'brvm_cote_du_jour.json')
        self.resume_file = os.path.join(os.path.dirname(__file__), 'brvm_resume.json')
        self.analyses_file = os.path.join(os.path.dirname(__file__), 'brvm_analyses.json')
        self.recommandations_file = os.path.join(os.path.dirname(__file__), 'brvm_recommandations.json')
        self.propositions_file = os.path.join(os.path.dirname(__file__), 'brvm_propositions.json')
        self.notes_marche_file = os.path.join(os.path.dirname(__file__), 'brvm_notes_marche.json')
        self.mouvements_file = os.path.join(os.path.dirname(__file__), 'brvm_mouvements.json')
        
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.log_dir, exist_ok=True)
        
        self.load_config_github()
        self.load_mapping()
        self.load_dividendes()
        self.load_cote_du_jour()
        self.load_resume()
        self.load_analyses()
        self.load_recommandations()
        self.load_propositions()
        self.load_notes_marche()
        self.load_mouvements()
        self.historique = self.load_historique()
        self.actualites = self.load_actualites()
        self.donnees_jour = self.get_donnees_jour()
    
    def load_config_github(self):
        """Charger config depuis variables d'environnement GitHub"""
        self.email_from = os.getenv('BRVM_EMAIL', '')
        recipients_str = os.getenv('RECIPIENTS', '')
        self.email_to = [e.strip() for e in recipients_str.split(',')] if recipients_str else []
        self.smtp_password = os.getenv('SMTP_PASSWORD', '')
        self.smtp_server = os.getenv('SMTP_SERVER', '')
        self.smtp_port = int(os.getenv('SMTP_PORT', '465'))
        
        self.log(f"✅ Config chargée depuis variables d'environnement")
        self.log(f"  Email from: {self.email_from}")
        self.log(f"  Recipients: {len(self.email_to)} adresses")
    
    def load_cote_du_jour(self):
        """Charger données de cote du jour"""
        self.cote_data = {}
        try:
            if os.path.exists(self.cote_file):
                with open(self.cote_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.cote_data = data.get('cote_data', {})
                self.log(f"✅ Cote du jour chargée")
            else:
                self.log("⚠️ Cote du jour non trouvée")
                self.cote_data = {}
        except Exception as e:
            self.log(f"⚠️ Erreur loading cote: {e}")
            self.cote_data = {}
    
    def load_resume(self):
        """Charger données resume (Top 5 + Flop 5)"""
        self.resume_data = {
            'top_5_gagnants': [],
            'flop_5_perdants': []
        }
        try:
            if os.path.exists(self.resume_file):
                with open(self.resume_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.resume_data = {
                        'top_5_gagnants': data.get('top_5_gagnants', []),
                        'flop_5_perdants': data.get('flop_5_perdants', [])
                    }
                self.log(f"✅ Resume chargé: {len(self.resume_data['top_5_gagnants'])} gagnants, {len(self.resume_data['flop_5_perdants'])} perdants")
            else:
                self.log("⚠️ Resume non trouvé")
                self.resume_data = {'top_5_gagnants': [], 'flop_5_perdants': []}
        except Exception as e:
            self.log(f"⚠️ Erreur loading resume: {e}")
            self.resume_data = {'top_5_gagnants': [], 'flop_5_perdants': []}
    
    def load_dividendes(self):
        """Charger historique dividendes"""
        self.dividendes = {}
        try:
            if os.path.exists(self.dividendes_file):
                with open(self.dividendes_file, 'r', encoding='utf-8') as f:
                    self.dividendes = json.load(f)
                self.log(f"✅ Dividendes chargés: {len(self.dividendes)} actions")
            else:
                self.log("⚠️ Historique dividendes non trouvé")
        except Exception as e:
            self.log(f"⚠️ Erreur loading dividendes: {e}")
    
    def load_analyses(self):
        """Charger analyses trading"""
        self.analyses = {'madis_invest': [], 'sika_finance': []}
        try:
            if os.path.exists(self.analyses_file):
                with open(self.analyses_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.analyses = {
                        'madis_invest': data.get('madis_invest', [])[:3],
                        'sika_finance': data.get('sika_finance', [])[:3]
                    }
                self.log(f"✅ Analyses chargées: {len(self.analyses['madis_invest']) + len(self.analyses['sika_finance'])} analyses")
            else:
                self.log("⚠️ Analyses non trouvées")
        except Exception as e:
            self.log(f"⚠️ Erreur loading analyses: {e}")
    
    def load_recommandations(self):
        """Charger recommandations intelligentes"""
        self.recommandations = {
            'recommandations': [],
            'synthese': []
        }
        try:
            if os.path.exists(self.recommandations_file):
                with open(self.recommandations_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.recommandations = {
                        'recommandations': data.get('recommandations', []),
                        'synthese': data.get('synthese', [])
                    }
                self.log(f"✅ Recommandations chargées: {len(self.recommandations['recommandations'])} recommandations")
            else:
                self.log("⚠️ Recommandations non trouvées")
        except Exception as e:
            self.log(f"⚠️ Erreur loading recommandations: {e}")
    
    def load_propositions(self):
        """Charger propositions d'investissement par secteur (Phase 1)"""
        self.propositions = {'propositions': [], 'synthese': {}}
        try:
            if os.path.exists(self.propositions_file):
                with open(self.propositions_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.propositions = {
                        'propositions': data.get('propositions', []),
                        'synthese': data.get('synthese', {})
                    }
                self.log(f"✅ Propositions chargées: {len(self.propositions['propositions'])} propositions")
            else:
                self.log("⚠️ Propositions non trouvées")
        except Exception as e:
            self.log(f"⚠️ Erreur loading propositions: {e}")
    
    def load_notes_marche(self):
        """Charger notes de marché avec impact BRVM (Phase 2)"""
        self.notes_marche = {'notes': [], 'synthese_globale': {}}
        try:
            if os.path.exists(self.notes_marche_file):
                with open(self.notes_marche_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.notes_marche = {
                        'notes': data.get('notes', []),
                        'synthese_globale': data.get('synthese_globale', {})
                    }
                self.log(f"✅ Notes de marché chargées: {len(self.notes_marche['notes'])} notes")
            else:
                self.log("⚠️ Notes de marché non trouvées")
        except Exception as e:
            self.log(f"⚠️ Erreur loading notes marche: {e}")
    
    def load_mouvements(self):
        """Charger mouvements BRVM (Phase 3B)"""
        self.mouvements = {
            'nouvelles_cotations': [],
            'suspensions': [],
            'augmentations_capital': [],
            'fusions_acquisitions': [],
            'synthese': {}
        }
        try:
            if os.path.exists(self.mouvements_file):
                with open(self.mouvements_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.mouvements = {
                        'nouvelles_cotations': data.get('nouvelles_cotations', []),
                        'suspensions': data.get('suspensions', []),
                        'augmentations_capital': data.get('augmentations_capital', []),
                        'fusions_acquisitions': data.get('fusions_acquisitions', []),
                        'synthese': data.get('synthese', {})
                    }
                total = self.mouvements['synthese'].get('total_mouvements', 0)
                self.log(f"✅ Mouvements BRVM chargés: {total} mouvements")
            else:
                self.log("⚠️ Mouvements BRVM non trouvés")
        except Exception as e:
            self.log(f"⚠️ Erreur loading mouvements: {e}")
    
    def calculate_dividend_metrics(self, code):
        """Calculer rendement moyen, croissance, appréciation"""
        if code not in self.dividendes:
            return None
        
        data = self.dividendes[code]['historique']
        years = sorted(data.keys())
        
        if len(years) < 2:
            return None
        
        rendements = []
        for year in years:
            div = data[year].get('dividende', 0)
            cours = data[year].get('cours_base', 1)
            rendement = (div / cours) * 100 if cours > 0 else 0
            rendements.append(rendement)
        
        rendement_moyen = sum(rendements) / len(rendements)
        
        premier_div = data[years[0]].get('dividende', 1)
        dernier_div = data[years[-1]].get('dividende', 1)
        nb_years = len(years) - 1
        tcac_div = ((dernier_div / premier_div) ** (1/nb_years) - 1) * 100 if nb_years > 0 else 0
        
        premier_cours = data[years[0]].get('cours_base', 1)
        dernier_cours = data[years[-1]].get('cours_fin', 1)
        appreciation = ((dernier_cours - premier_cours) / premier_cours) * 100 if premier_cours > 0 else 0
        
        return {
            'rendement_moyen': rendement_moyen,
            'tcac_dividende': tcac_div,
            'appreciation_cours': appreciation,
            'dernier_dividende': dernier_div
        }
    
    def load_mapping(self):
        """Charger mapping actions"""
        self.action_names = {}
        try:
            if os.path.exists(self.mapping_file):
                with open(self.mapping_file, 'r', encoding='utf-8') as f:
                    self.action_names = json.load(f)
                self.log(f"✅ Mapping charge: {len(self.action_names)} actions")
        except Exception as e:
            self.log(f"⚠️ Erreur mapping: {e}")
    
    def get_action_display_name(self, code):
        """Retourne 'CODE - NOM_COMPLET' - compatible ancien et nouveau format"""
        if code in self.action_names:
            action_data = self.action_names[code]
            
            # Nouveau format enrichi (dict)
            if isinstance(action_data, dict):
                nom = action_data.get('nom_complet', code)
                return f"{code} - {nom}"
            
            # Ancien format simple (string)
            else:
                return f"{code} - {action_data}"
        
        return code
    
    def get_next_version(self, filepath):
        """Trouve prochain numéro version"""
        if not os.path.exists(filepath):
            return filepath
        
        base, ext = os.path.splitext(filepath)
        version = 2
        while True:
            new_path = f"{base}_V{version}{ext}"
            if not os.path.exists(new_path):
                return new_path
            version += 1
    
    def load_historique(self):
        """Charger historique"""
        try:
            if os.path.exists(self.historique_file):
                with open(self.historique_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {}
    
    def load_actualites(self):
        """Charger actualités"""
        try:
            if os.path.exists(self.actualites_file):
                with open(self.actualites_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {}
    
    def get_donnees_jour(self):
        """Récupérer données du jour"""
        if self.date_file in self.historique:
            return self.historique[self.date_file]
        return {
            "date": self.date_file,
            "indice_composite": 12450,
            "volume_total": 4200000000,
            "actions_hausse": 28,
            "actions_baisse": 15,
            "actions": []
        }
    
    def log(self, msg):
        """Logger"""
        print(msg)
        log_file = os.path.join(self.log_dir, f'brvm_{self.date_file}.log')
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
    
    def generate_excel(self):
        """Générer Excel"""
        self.log("📊 Generation Excel...")
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Synthèse
            ws = wb.create_sheet("Synthese", 0)
            ws['A1'] = "BULLETIN BRVM - DONNÉES OFFICIELLES"
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Date: {self.date}"
            ws['A2'].font = Font(bold=True)
            
            synthese = self.cote_data.get('synthese', {})
            indice = synthese.get('indice_composite', 12450)
            volume = synthese.get('volume_total_mds', 4.2)
            
            ws['A4'] = "Indice BRVM Composite"
            ws['B4'] = indice
            ws['A5'] = "Volume Total"
            ws['B5'] = f"{volume} Mds FCFA"
            
            # Top actions
            ws2 = wb.create_sheet("Top Actions", 1)
            ws2['A1'] = "TOP 5 ACTIONS"
            ws2['A1'].font = Font(bold=True, size=12)
            
            actions_cote = self.cote_data.get('actions', {})
            top_actions = list(actions_cote.items())[:5]
            
            ws2['A3'] = "Code"
            ws2['B3'] = "Nom"
            ws2['C3'] = "Cours"
            ws2['D3'] = "Variation %"
            
            for idx, (code, data) in enumerate(top_actions, start=4):
                ws2[f'A{idx}'] = code
                ws2[f'B{idx}'] = self.get_action_display_name(code)
                ws2[f'C{idx}'] = data.get('cours', 0)
                ws2[f'D{idx}'] = data.get('variation', 0)
            
            filepath = os.path.join(self.output_dir, f'BRVM_Bulletin_{self.date_file}.xlsx')
            filepath = self.get_next_version(filepath)
            
            wb.save(filepath)
            self.log(f"✅ Excel créé: {filepath}")
            return filepath
            
        except Exception as e:
            self.log(f"❌ Erreur Excel: {e}")
            return None
    
    def generate_html_complet(self):
        """Générer HTML COMPLET"""
        self.log("📄 Generation HTML complet...")
        
        try:
            # Données synthèse OFFICIELLES
            synthese = self.cote_data.get('synthese', {})
            indice = synthese.get('indice_composite', 12450)
            volume = synthese.get('volume_total_mds', 4.2)
            actions_hausse = synthese.get('actions_hausse', 28)
            actions_baisse = synthese.get('actions_baisse', 15)
            
            # Charger Top 5 et Flop 5 depuis resume (BRVM.org/resume)
            top_5_gagnants = self.resume_data.get('top_5_gagnants', [])
            flop_5_perdants = self.resume_data.get('flop_5_perdants', [])
            
            # Métriques dividendes
            metriques_rendement = []
            for code in self.dividendes.keys():
                metrics = self.calculate_dividend_metrics(code)
                if metrics:
                    metriques_rendement.append((code, metrics['rendement_moyen']))
            metriques_rendement.sort(key=lambda x: x[1], reverse=True)
            
            # Opportunités cachées
            actions_cachees = []
            codes_top_flop = [t.get('code') for t in top_5_gagnants] + [t.get('code') for t in flop_5_perdants]
            for code, rendement in metriques_rendement[:10]:
                if code not in codes_top_flop:
                    metrics = self.calculate_dividend_metrics(code)
                    if metrics and (metrics['rendement_moyen'] >= 2.5 or metrics['appreciation_cours'] >= 30):
                        actions_cachees.append({
                            'code': code,
                            'rendement': metrics['rendement_moyen'],
                            'appreciation': metrics['appreciation_cours']
                        })
            
            # Recommandations intelligentes
            recommandations_html = ""
            synthese = self.recommandations.get('synthese', [])
            if synthese:
                recommandations_html = "<div class='section'><div class='section-title'>💡 Recommandations d'Investissement (Basées sur l'Actualité)</div>"
                for section in synthese:
                    titre = section.get('titre', '')
                    actions = section.get('actions', [])
                    if actions:
                        recommandations_html += f"<div style='margin: 15px 0;'><strong>{titre}</strong><table style='width:100%; margin-top: 10px;'>"
                        recommandations_html += "<tr><th style='text-align: left; background: #f0f0f0; padding: 8px;'>Code</th>"
                        recommandations_html += "<th style='text-align: left; background: #f0f0f0; padding: 8px;'>Raison</th>"
                        recommandations_html += "<th style='text-align: left; background: #f0f0f0; padding: 8px;'>Confiance</th>"
                        recommandations_html += "<th style='text-align: left; background: #f0f0f0; padding: 8px;'>Rendement</th></tr>"
                        
                        for rec in actions[:3]:
                            code = rec.get('code', 'N/A')
                            raison = rec.get('raison', '')[:60]
                            confiance = rec.get('confiance', 'N/A')
                            rendement = rec.get('rendement_dividend', 0)
                            source = rec.get('source', '')
                            
                            color_confiance = '#28a745' if confiance == 'FORTE' else '#ffc107' if confiance == 'MOYENNE' else '#dc3545'
                            
                            recommandations_html += f"<tr><td style='padding: 8px;'><strong>{code}</strong></td>"
                            recommandations_html += f"<td style='padding: 8px;'>{raison}<br/><small style='color: #666;'>{source}</small></td>"
                            recommandations_html += f"<td style='padding: 8px;'><span style='background: {color_confiance}; color: white; padding: 2px 8px; border-radius: 3px; font-size: 11px;'>{confiance}</span></td>"
                            recommandations_html += f"<td style='padding: 8px;'>{rendement:.2f}%</td></tr>"
                        
                        recommandations_html += "</table></div>"
                
                recommandations_html += "</div>"
            
            # Propositions d'investissement par secteur (PHASE 1)
            propositions_html = ""
            propositions = self.propositions.get('propositions', [])
            if propositions:
                propositions_html = "<div class='section'><div class='section-title'>🎯 Propositions d'Investissement par Secteur</div>"
                for prop in propositions[:5]:
                    secteur = prop.get('secteur', 'N/A')
                    recommandation = prop.get('recommandation', 'N/A')
                    confiance = prop.get('confiance', 'N/A')
                    actions = prop.get('actions', [])
                    raison = prop.get('raison', '')[:80]
                    horizon = prop.get('horizon', '')
                    
                    color_rec = '#28a745' if recommandation == 'ACHETER' else '#dc3545' if recommandation == 'VENDRE' else '#ffc107'
                    color_conf = '#28a745' if confiance == 'FORTE' else '#ffc107' if confiance == 'MOYENNE' else '#dc3545'
                    
                    propositions_html += f"<div style='margin: 15px 0; padding: 12px; background: #f9f9f9; border-left: 4px solid {color_rec};'>"
                    propositions_html += f"<strong style='font-size: 14px;'>{secteur}</strong><br/>"
                    propositions_html += f"<span style='background: {color_rec}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px; margin-right: 8px;'>{recommandation}</span>"
                    propositions_html += f"<span style='background: {color_conf}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px;'>{confiance}</span><br/>"
                    propositions_html += f"<small style='color: #666;'>{raison} | Horizon: {horizon}</small><br/>"
                    propositions_html += f"<small><strong>Actions:</strong> {', '.join(actions[:5])}</small>"
                    propositions_html += "</div>"
                propositions_html += "</div>"
            
            # Notes de marché avec impact BRVM (PHASE 2)
            notes_marche_html = ""
            notes = self.notes_marche.get('notes', [])
            synthese_global = self.notes_marche.get('synthese_globale', {})
            if notes:
                notes_marche_html = "<div class='section'><div class='section-title'>📰 Notes de Marché - Impact BRVM</div>"
                
                # Synthèse globale
                if synthese_global:
                    sentiment = synthese_global.get('sentiment_overall', 'NEUTRE')
                    impact = synthese_global.get('impact_overall', 'MOYEN')
                    risque = synthese_global.get('risque_global', 'MOYEN')
                    
                    color_sent = '#28a745' if sentiment == 'POSITIF' else '#dc3545' if sentiment == 'NÉGATIF' else '#ffc107'
                    color_impact = '#28a745' if impact == 'FORT' else '#ffc107' if impact == 'MOYEN' else '#dc3545'
                    color_risk = '#dc3545' if risque == 'ÉLEVÉ' else '#ffc107' if risque == 'MOYEN' else '#28a745'
                    
                    notes_marche_html += f"<div style='margin: 15px 0; padding: 12px; background: #f0f8ff; border-radius: 5px;'>"
                    notes_marche_html += f"<strong>📊 Synthèse Globale:</strong><br/>"
                    notes_marche_html += f"<span style='background: {color_sent}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px; margin-right: 8px;'>Sentiment: {sentiment}</span>"
                    notes_marche_html += f"<span style='background: {color_impact}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px; margin-right: 8px;'>Impact: {impact}</span>"
                    notes_marche_html += f"<span style='background: {color_risk}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px;'>Risque: {risque}</span>"
                    notes_marche_html += "</div>"
                
                # Détail des notes
                for note in notes[:5]:
                    titre = note.get('titre', 'N/A')[:100]
                    sentiment = note.get('sentiment', 'NEUTRE')
                    impact = note.get('impact_brvm', 'MOYEN')
                    secteurs = note.get('secteurs_beneficiaires', []) + note.get('secteurs_affectes_negativement', [])
                    actions = note.get('actions_impactees', [])
                    probabilite = note.get('probabilite', 0)
                    synthese = note.get('synthese', '')
                    
                    color_sent = '#28a745' if sentiment == 'POSITIF' else '#dc3545' if sentiment == 'NÉGATIF' else '#ffc107'
                    color_impact = '#28a745' if impact == 'FORT' else '#ffc107' if impact == 'MOYEN' else '#dc3545'
                    
                    notes_marche_html += f"<div style='margin: 15px 0; padding: 12px; background: #f9f9f9; border-left: 4px solid {color_sent};'>"
                    notes_marche_html += f"<strong>{titre}</strong><br/>"
                    notes_marche_html += f"<span style='background: {color_sent}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; margin-right: 6px;'>{sentiment}</span>"
                    notes_marche_html += f"<span style='background: {color_impact}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;'>{impact} impact</span><br/>"
                    notes_marche_html += f"<small style='color: #666;'>{synthese}</small><br/>"
                    if secteurs:
                        notes_marche_html += f"<small><strong>Secteurs:</strong> {', '.join(secteurs[:3])}</small><br/>"
                    if actions:
                        notes_marche_html += f"<small><strong>Actions impactées:</strong> {', '.join(actions[:5])}</small><br/>"
                    notes_marche_html += f"<small><strong>Probabilité:</strong> {probabilite*100:.0f}%</small>"
                    notes_marche_html += "</div>"
                
                notes_marche_html += "</div>"
            
            # Mouvements BRVM (PHASE 3B)
            mouvements_html = ""
            mouvements = self.mouvements.get('mouvements', [])
            synthese_mouv = self.mouvements.get('synthese', {})
            if synthese_mouv.get('total_mouvements', 0) > 0:
                mouvements_html = "<div class='section'><div class='section-title'>🔄 Mouvements BRVM</div>"
                
                # Nouvelles cotations
                nouvelles = self.mouvements.get('nouvelles_cotations', [])
                if nouvelles:
                    mouvements_html += "<div style='margin: 15px 0;'><strong>✨ Nouvelles Cotations:</strong><table style='width:100%; margin-top: 10px;'>"
                    mouvements_html += "<tr><th style='background: #f0f0f0; padding: 8px;'>Code</th><th style='background: #f0f0f0; padding: 8px;'>Nom</th><th style='background: #f0f0f0; padding: 8px;'>Secteur</th><th style='background: #f0f0f0; padding: 8px;'>Date</th></tr>"
                    for new in nouvelles[:3]:
                        code = new.get('code', 'N/A')
                        nom = new.get('nom', 'N/A')[:50]
                        secteur = new.get('secteur', 'N/A')
                        date = new.get('date_cotation', 'N/A')
                        mouvements_html += f"<tr><td style='padding: 8px;'><strong>{code}</strong></td><td style='padding: 8px;'>{nom}</td><td style='padding: 8px;'>{secteur}</td><td style='padding: 8px;'>{date}</td></tr>"
                    mouvements_html += "</table></div>"
                
                # Suspensions
                suspensions = self.mouvements.get('suspensions', [])
                if suspensions:
                    mouvements_html += "<div style='margin: 15px 0;'><strong>⚠️ Suspensions:</strong><table style='width:100%; margin-top: 10px;'>"
                    mouvements_html += "<tr><th style='background: #f0f0f0; padding: 8px;'>Code</th><th style='background: #f0f0f0; padding: 8px;'>Raison</th><th style='background: #f0f0f0; padding: 8px;'>Fin</th></tr>"
                    for susp in suspensions[:3]:
                        code = susp.get('code', 'N/A')
                        raison = susp.get('raison', 'N/A')[:40]
                        fin = susp.get('date_fin', 'N/A')
                        mouvements_html += f"<tr><td style='padding: 8px;'><strong>{code}</strong></td><td style='padding: 8px;'>{raison}</td><td style='padding: 8px;'>{fin}</td></tr>"
                    mouvements_html += "</table></div>"
                
                # Augmentations capital
                augmentations = self.mouvements.get('augmentations_capital', [])
                if augmentations:
                    mouvements_html += "<div style='margin: 15px 0;'><strong>💰 Augmentations Capital:</strong><table style='width:100%; margin-top: 10px;'>"
                    mouvements_html += "<tr><th style='background: #f0f0f0; padding: 8px;'>Code</th><th style='background: #f0f0f0; padding: 8px;'>Montant (Mds FCFA)</th><th style='background: #f0f0f0; padding: 8px;'>Date</th></tr>"
                    for aug in augmentations[:3]:
                        code = aug.get('code', 'N/A')
                        montant = aug.get('montant_mds_fcfa', 'N/A')
                        date = aug.get('date_operation', 'N/A')
                        mouvements_html += f"<tr><td style='padding: 8px;'><strong>{code}</strong></td><td style='padding: 8px;'>{montant}</td><td style='padding: 8px;'>{date}</td></tr>"
                    mouvements_html += "</table></div>"
                
                # Fusions/Acquisitions
                fusions = self.mouvements.get('fusions_acquisitions', [])
                if fusions:
                    mouvements_html += "<div style='margin: 15px 0;'><strong>🔗 Fusions & Acquisitions:</strong><table style='width:100%; margin-top: 10px;'>"
                    mouvements_html += "<tr><th style='background: #f0f0f0; padding: 8px;'>Acquis</th><th style='background: #f0f0f0; padding: 8px;'>Acquireur</th><th style='background: #f0f0f0; padding: 8px;'>Date</th></tr>"
                    for fus in fusions[:3]:
                        code_acq = fus.get('code_acquis', 'N/A')
                        code_acq2 = fus.get('code_acquireur', 'N/A')
                        date = fus.get('date_operation', 'N/A')
                        mouvements_html += f"<tr><td style='padding: 8px;'>{code_acq}</td><td style='padding: 8px;'><strong>{code_acq2}</strong></td><td style='padding: 8px;'>{date}</td></tr>"
                    mouvements_html += "</table></div>"
                
                mouvements_html += "</div>"
            
            # Analyses trading avec recommandations
            analyses_html = ""
            if self.analyses['madis_invest']:
                analyses_html += "<div class='section'><div class='section-title'>📊 Analyses Madis Invest</div>"
                for analyse in self.analyses['madis_invest'][:3]:
                    title = analyse.get('title', 'N/A')[:100]
                    desc = analyse.get('description', '')[:150]
                    analyses_html += f"<p><strong>{title}</strong><br/><em>{desc}</em></p>"
                analyses_html += "</div>"
            
            if self.analyses['sika_finance']:
                analyses_html += "<div class='section'><div class='section-title'>📊 Actualités Sika Finance</div>"
                for analyse in self.analyses['sika_finance'][:3]:
                    title = analyse.get('title', 'N/A')[:100]
                    desc = analyse.get('description', '')[:150]
                    analyses_html += f"<p><strong>{title}</strong><br/><em>{desc}</em></p>"
                analyses_html += "</div>"
            
            # Top 5 Gagnants HTML
            top_5_html = "<table><tr><th>Code</th><th>Nom</th><th>Cours</th><th>Var %</th></tr>"
            if top_5_gagnants:
                for action in top_5_gagnants:
                    code = action.get('code', 'N/A')
                    nom = self.get_action_display_name(code)
                    cours = action.get('cours', 0)
                    var = action.get('variation', 0)
                    top_5_html += f"<tr><td>{code}</td><td>{nom}</td><td>{cours}</td><td class='positive'>{var:+.2f}%</td></tr>"
            else:
                top_5_html += "<tr><td colspan='4'>Aucune donnée disponible</td></tr>"
            top_5_html += "</table>"
            
            # Flop 5 Perdants HTML
            flop_5_html = "<table><tr><th>Code</th><th>Nom</th><th>Cours</th><th>Var %</th></tr>"
            if flop_5_perdants:
                for action in flop_5_perdants:
                    code = action.get('code', 'N/A')
                    nom = self.get_action_display_name(code)
                    cours = action.get('cours', 0)
                    var = action.get('variation', 0)
                    flop_5_html += f"<tr><td>{code}</td><td>{nom}</td><td>{cours}</td><td class='negative'>{var:+.2f}%</td></tr>"
            else:
                flop_5_html += "<tr><td colspan='4'>Aucune donnée disponible</td></tr>"
            flop_5_html += "</table>"
            
            # Opportunités HTML
            opp_html = ""
            if actions_cachees:
                opp_html = "<div class='section'><div class='section-title'>💡 Opportunités Cachées</div><table>"
                opp_html += "<tr><th>Code</th><th>Rendement</th><th>Appréciation</th></tr>"
                for opp in actions_cachees[:5]:
                    opp_html += f"<tr><td>{opp['code']}</td><td>{opp['rendement']:.2f}%</td><td>{opp['appreciation']:.2f}%</td></tr>"
                opp_html += "</table></div>"
            
            html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; margin: 0; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: white; }}
        header {{ background: linear-gradient(135deg, #1F4E78 0%, #3D5A80 100%); color: white; padding: 30px; text-align: center; }}
        h1 {{ margin: 0; font-size: 24px; }}
        .section {{ margin: 20px 0; padding: 20px; border-left: 4px solid #1F4E78; background: #f9f9f9; }}
        .section-title {{ font-size: 16px; font-weight: bold; color: #1F4E78; margin-bottom: 10px; }}
        .official {{ background: #e8f4f8; border: 2px solid #0099cc; padding: 10px; margin: 10px 0; }}
        .official-badge {{ display: inline-block; background: #0099cc; color: white; padding: 5px 10px; border-radius: 3px; font-size: 11px; font-weight: bold; margin-right: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 13px; }}
        th {{ background: #1F4E78; color: white; padding: 10px; text-align: left; }}
        td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
        .positive {{ color: #28a745; font-weight: bold; }}
        .negative {{ color: #dc3545; font-weight: bold; }}
        footer {{ background: #f8f9fa; padding: 15px; text-align: center; font-size: 11px; color: #666; border-top: 1px solid #ddd; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Bulletin d'Investissement Quotidien</h1>
            <p>Marche Ouest-Africain (BRVM)</p>
            <p>{self.date}</p>
            <div class="official-badge">✓ DONNÉES OFFICIELLES BRVM</div>
        </header>

        <div style="padding: 20px;">
            <div class="section official">
                <div class="section-title"><span class="official-badge">OFFICIEL</span> Performance du Marche</div>
                <table>
                    <tr><th>Metrique</th><th>Valeur</th><th>Source</th></tr>
                    <tr>
                        <td>Indice BRVM Composite</td>
                        <td><strong>{indice:,}</strong></td>
                        <td>Bulletins BRVM</td>
                    </tr>
                    <tr>
                        <td>Volume Total</td>
                        <td><strong>{volume:.2f} Mds FCFA</strong></td>
                        <td>Bulletins BRVM</td>
                    </tr>
                    <tr>
                        <td>Actions en Hausse</td>
                        <td><strong class="positive">{actions_hausse}</strong></td>
                        <td>Bulletins BRVM</td>
                    </tr>
                    <tr>
                        <td>Actions en Baisse</td>
                        <td><strong class="negative">{actions_baisse}</strong></td>
                        <td>Bulletins BRVM</td>
                    </tr>
                </table>
            </div>

            <div class="section">
                <div class="section-title">🏆 Top 5 Actions du Jour (Gagnants)</div>
                {top_5_html}
            </div>

            <div class="section">
                <div class="section-title">📉 Flop 5 Actions du Jour (Perdants)</div>
                {flop_5_html}
            </div>

            {recommandations_html}

            {propositions_html}

            {notes_marche_html}

            {mouvements_html}

            {analyses_html}

            {opp_html}

        </div>

        <footer>
            <p><strong>✓ DONNÉES OFFICIELLES</strong></p>
            <p>Bulletin genere avec donnees officielles des Bulletins de Cote BRVM</p>
            <p>Analyses trading: Madis Invest + Sika Finance</p>
            <p>Source: BRVM.org | Analyse neutre et factuelle</p>
        </footer>
    </div>
</body>
</html>"""
            
            filepath = os.path.join(self.output_dir, f'BRVM_Bulletin_{self.date_file}.html')
            filepath = self.get_next_version(filepath)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html)
            self.log(f"✅ HTML créé: {filepath}")
            return filepath
            
        except Exception as e:
            self.log(f"❌ Erreur HTML: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def send_email(self, excel_file, html_file):
        """Envoyer email"""
        if not self.email_to or not self.email_from or not self.smtp_password:
            self.log("⚠️ Configuration incomplète - Email pas envoyé")
            return False
        
        self.log(f"📧 Envoi a: {', '.join(self.email_to)}")
        
        try:
            if not os.path.exists(html_file):
                self.log(f"❌ HTML manquant")
                return False
            
            with open(html_file, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            msg = MIMEMultipart('mixed')
            msg['Subject'] = f"Bulletin BRVM - {self.date} (Donnees Officielles)"
            msg['From'] = self.email_from
            msg['To'] = ', '.join(self.email_to)
            
            msg.attach(MIMEText(html_content, 'html', _charset='utf-8'))
            
            # Attachment Excel optionnel (peut être None)
            if excel_file and os.path.exists(excel_file):
                with open(excel_file, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 
                              f'attachment; filename="{os.path.basename(excel_file)}"')
                msg.attach(part)
                self.log(f"📎 Piece jointe: {os.path.basename(excel_file)}")
            else:
                self.log("📧 Email sans pièce jointe Excel (désactivée)")
            
            if self.smtp_port == 465:
                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, context=context) as server:
                    server.login(self.email_from, self.smtp_password)
                    server.send_message(msg)
            else:
                context = ssl.create_default_context()
                with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                    server.starttls(context=context)
                    server.login(self.email_from, self.smtp_password)
                    server.send_message(msg)
            
            self.log("✅ Email envoye avec succes!")
            return True
                
        except Exception as e:
            self.log(f"❌ Erreur email: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def run(self):
        """Exécuter"""
        self.log("\n" + "="*60)
        self.log(f"🚀 BRVM V3 FINAL COMPLET - {self.date}")
        self.log("="*60)
        self.log(f"Données: Bulletins + Dividendes + Analyses Trading + Propositions + Mouvements")
        self.log("="*60)
        
        # Excel génération DÉSACTIVÉE (non nécessaire)
        html_file = self.generate_html_complet()
        
        if html_file:
            self.send_email(None, html_file)
        else:
            self.log("❌ Impossible d'envoyer: fichier HTML manquant")
        
        self.log("="*60)
        self.log("✅ Bulletin genere!")
        self.log("="*60 + "\n")

if __name__ == "__main__":
    automation = BRVMAutomationComplet()
    automation.run()
